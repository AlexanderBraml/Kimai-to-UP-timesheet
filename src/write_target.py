import datetime
from typing import Tuple, List

from openpyxl import load_workbook

from src.get_kimai_data import read_csv

date_index = (5, 12)

time_row = 12

date_col_1 = 1
start_hour_col_1 = 2
start_min_col_1 = 4
end_hour_col_1 = 7
end_min_col_1 = 9

max_rows = 31


def write_to_file(target_template: str, target_file: str,
                  times: List[Tuple[datetime.datetime, datetime.datetime]]) -> None:
    assert len(times) > 0

    workbook = load_workbook(filename=target_template)
    sheet = workbook.active

    sheet.cell(*date_index).value = times[0][0].replace(day=1)

    for idx, time in enumerate(times):
        row = time_row + idx
        sheet.cell(row, date_col_1).value = time[0].strftime("%d.%m.%Y")

        sheet.cell(row, start_hour_col_1).value = time[0].hour
        sheet.cell(row, start_min_col_1).value = time[0].minute
        sheet.cell(row, end_hour_col_1).value = time[1].hour
        sheet.cell(row, end_min_col_1).value = time[1].minute

    workbook.save(target_file)
    workbook.close()


if __name__ == '__main__':
    times = read_csv('../templates/source.csv')
    write_to_file('../templates/target.xlsx', '../templates/April2023.xlsx', times)
